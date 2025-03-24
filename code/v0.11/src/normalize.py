import os
import numpy as np
import nibabel as nib
import scipy.io as sio
from scipy.ndimage import center_of_mass, shift
from skimage.measure import regionprops
import openpyxl

# --- License information (not executable code) ---
# This script is part of the aVP-Toolbox v0.11 - 2023 software.
# ... (rest of the license text)

# --- Setup ---
curwd = os.getcwd()
with open(os.path.join(curwd, 'ONcontrol.txt'), 'r') as f:
    StudyPath = f.readline().strip()

inPath = os.path.join(StudyPath, 'data', 'proc')
outImPath = os.path.join(StudyPath, 'data', 'proc')
outResPath = os.path.join(StudyPath, 'results')

# --- Add toolboxes (adjust paths as needed) ---
# Note: The MATLAB toolboxes need to be converted to Python equivalents.
#       This code assumes that you have equivalent functions for:
#       - load_nifti (e.g., using nibabel)
#       - save_nifti (e.g., using nibabel)
#       - xlwrite (e.g., using openpyxl)
#       - regionprops (scikit-image)
#       - circshift (scipy.ndimage.shift)

# Example using nibabel:
# import nibabel as nib

# Example using openpyxl:
# from openpyxl import Workbook

# --- Java libraries for xlwrite (not needed with openpyxl) ---
# (These are not needed if you use a pure Python library like openpyxl)

# --- Output files and headers ---
DataFile = os.path.join(outResPath, 'aVP_slice_data.xlsx')
HoleFile = os.path.join(outResPath, 'log_check_hole.xlsx')
RangeFile = os.path.join(outResPath, 'log_check_range.xlsx')
LenStretchFile = os.path.join(outResPath, 'aVP_section_CSA_length.xlsx')

tablabels = ['curr_sli_yz', 'orig_sli_yz', 'point_y', 'point_z', 'circshift_y', 'circshift_z', 'dist', 'int_dist_x10', 'len_on', 'tot_len', 'mMax', 'save_len', 'CSArea', 'Eccent', 'MajAxis', 'MinAxis', 'AvgCSA']
stretxt = ['Subject', 'ONsection', 'side', 'TotLength', 'OT_length', 'OC_length', 'iCran_length', 'iCan_length', 'iOrb_length', 'OT_CSA', 'OC_CSA', 'iCran_CSA', 'iCan_CSA', 'iOrb_CSA', 'SegmCode 1', 'SegmCode 2', 'SegmCode 3', 'SegmCode 4', 'SegmCode 5']

rangetxt = ['Subject', 'ONsection', 'side', 'Slice...']
# xlRange = 'A1'  # Not needed with openpyxl
write_excel(RangeFile, [rangetxt], 'Sheet 1', 'A1') # Using a helper function (see below)

holetxt = ['Subject', 'ONsection', 'side', 'Slice...']
write_excel(HoleFile, [holetxt], 'Sheet 1', 'A1')

Lin4image = '_linearize_4bc.nii.gz'
Lin4mat = '_linearize_4bc.mat'

Norm4image = '_normalized_4bc.nii.gz'
Norm4mat = '_normalized_4bc.mat'

fullNorm4image = '_full_normalized_4bc.nii.gz'

sheet = 1
sides = ['l', 'r']
isbj = 0
tablelength = []

maxNslices = 2500
ismask = 0

with open(os.path.join(StudyPath, 'data', 'sbj.list'), 'r') as f:
    subject_list = [line.strip() for line in f]

# --- Helper function for writing to Excel (using openpyxl) ---
def write_excel(filename, data, sheetname, start_cell):
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheetname)  # Create sheet if it doesn't exist.
    row_num = openpyxl.utils.cell.row_index_from_cell(start_cell)
    col_num = openpyxl.utils.cell.column_index_from_cell(start_cell)
    for row_index, row in enumerate(data):
        for col_index, cell_value in enumerate(row):
            ws.cell(row=row_num + row_index, column=col_num + col_index, value=cell_value)
    wb.save(filename)


results = []

for s, sbj in enumerate(subject_list):
    isbj += 1
    for side in sides:
        cc_value = [] # Initialize for each subject and side.
        fname = os.path.join(inPath, sbj, f'on_{side}')

        if not os.path.exists(fname + '.nii.gz'):
            print(f'Could not find: {fname}.nii.gz')
            continue  # Use continue instead of return to proceed with the next subject/side.

        aswap = nib.load(fname + '.nii.gz')
        aa = aswap.copy() # Correctly copy the NIfTI image object
        aa.img = np.flip(aswap.get_fdata(), axis=0) # Use get_fdata() and np.flip()
        oo = aa.copy()

        x_dim = aa.img.shape[0]
        z_dim = aa.img.shape[2]
        y_dim = aa.img.shape[1]

        x_resolution = aa.header['pixdim'][1]
        z_resolution = aa.header['pixdim'][3]
        y_resolution = aa.header['pixdim'][2]

        image_center = [z_dim / 2, x_dim / 2]
        active_slice_counter = 0
        segment_type = 0

        table = []

        current_max_value = 0

        for y in range(y_dim):  # Python indexing starts from 0
            selected_y_slice = aa.img[:, y, :]

            max_voxel_value = np.max(selected_y_slice)

            if max_voxel_value > 0:
                if active_slice_counter == 0:
                    segment_type = 1
                    current_max_value = max_voxel_value
                elif current_max_value != max_voxel_value:
                    segment_type += 1

                active_slice_counter += 1

                cc_value.append({}) # Append a dictionary to the list
                cc_value[-1]['current_slice_yz'] = active_slice_counter
                cc_value[-1]['original_slice_yz'] = y + 1 # Add 1 for MATLAB indexing
                cc_value[-1]['flagC'] = segment_type
                cc_value[-1]['mm'] = max_voxel_value

                binarized_slice = selected_y_slice > 0
                st = regionprops(binarized_slice.astype(int), 'centroid') # regionprops needs integer input
                area0 = regionprops(binarized_slice.astype(int), 'area')

                temp_centroids = np.array([prop.centroid for prop in st])
                centroids = temp_centroids[0]

                slice_properties = regionprops(binarized_slice.astype(int), 'centroid', 'Area', 'MajorAxisLength', 'MinorAxisLength', 'Eccentricity')

                cc_value[-1]['majaxis'] = slice_properties[0].MajorAxisLength * x_resolution
                cc_value[-1]['minaxis'] = slice_properties[0].MinorAxisLength * z_resolution
                cc_value[-1]['area'] = slice_properties[0].Area * x_resolution * z_resolution
                cc_value[-1]['eccent'] = slice_properties[0].Eccentricity


                cc = shift(selected_y_slice, [image_center[1] - round(centroids[1]), image_center[0] - round(centroids[0])]) # Use shift for circshift
                cc_value[-1]['circshift'] = [image_center[1] - round(centroids[1]), image_center[0] - round(centroids[0])]

                if ismask == 1:
                    oo.img[:, y, :] = np.sign